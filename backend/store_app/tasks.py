import os
from datetime import datetime
from decimal import Decimal
from django.conf import settings
from django.db.models import Sum
from celery import shared_task
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from store_app.models import Invoice, InvoiceItem

@shared_task
def generate_monthly_gst_report_task(month, year, admin_email=None):
    """
    Asynchronously compile monthly GST collections and output a styled Excel file.
    """
    # 1. Query Invoices for the target period
    invoices = Invoice.objects.filter(created_at__year=year, created_at__month=month).order_by('created_at')

    # Create Workbook
    wb = Workbook()
    
    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "GST Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Styling helpers
    font_title = Font(name='Calibri', size=16, bold=True, color='0F172A')
    font_section = Font(name='Calibri', size=12, bold=True, color='10B981')
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_bold = Font(name='Calibri', size=11, bold=True)
    font_regular = Font(name='Calibri', size=11)
    
    fill_header = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
    fill_zebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Headers
    ws_summary.merge_cells('A1:F1')
    ws_summary['A1'] = f"Shivam Kirana Store - GST Audit Summary ({month:02d}/{year})"
    ws_summary['A1'].font = font_title
    ws_summary['A1'].alignment = align_center
    
    ws_summary.append([]) # Blank row
    ws_summary.append([]) # Blank row
    
    ws_summary.append(["Tax Slab (%)", "Taxable Value (₹)", "CGST (₹)", "SGST (₹)", "Total Tax (₹)", "Total Sales (₹)"])
    for col_idx in range(1, 7):
        cell = ws_summary.cell(row=4, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    slabs = [Decimal('0.00'), Decimal('5.00'), Decimal('12.00'), Decimal('18.00'), Decimal('28.00')]
    invoice_items = InvoiceItem.objects.filter(invoice__in=invoices)
    
    row_num = 5
    total_taxable_all = Decimal('0.00')
    total_cgst_all = Decimal('0.00')
    total_sgst_all = Decimal('0.00')
    total_sales_all = Decimal('0.00')
    
    for slab in slabs:
        items = invoice_items.filter(gst_rate=slab)
        agg = items.aggregate(
            total_sales=Sum('total_amount'),
            cgst=Sum('cgst_amount'),
            sgst=Sum('sgst_amount')
        )
        slab_inclusive = agg['total_sales'] or Decimal('0.00')
        slab_cgst = agg['cgst'] or Decimal('0.00')
        slab_sgst = agg['sgst'] or Decimal('0.00')
        slab_taxable = slab_inclusive - (slab_cgst + slab_sgst)
        slab_tax = slab_cgst + slab_sgst
        
        ws_summary.append([
            f"{slab}%",
            float(slab_taxable),
            float(slab_cgst),
            float(slab_sgst),
            float(slab_tax),
            float(slab_inclusive)
        ])
        
        for col_idx in range(1, 7):
            cell = ws_summary.cell(row=row_num, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            cell.alignment = align_right if col_idx > 1 else align_center
            if row_num % 2 == 0:
                cell.fill = fill_zebra
                
        total_taxable_all += slab_taxable
        total_cgst_all += slab_cgst
        total_sgst_all += slab_sgst
        total_sales_all += slab_inclusive
        row_num += 1
        
    # Add Total Row
    ws_summary.append([
        "Total",
        float(total_taxable_all),
        float(total_cgst_all),
        float(total_sgst_all),
        float(total_cgst_all + total_sgst_all),
        float(total_sales_all)
    ])
    for col_idx in range(1, 7):
        cell = ws_summary.cell(row=row_num, column=col_idx)
        cell.font = font_bold
        cell.border = thin_border
        cell.alignment = align_right if col_idx > 1 else align_center
        cell.fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
        
    # --- Sheet 2: Invoices ---
    ws_invoices = wb.create_sheet(title="Invoice Ledger")
    ws_invoices.views.sheetView[0].showGridLines = True
    
    ws_invoices.append(["Invoice Number", "Date", "Customer", "Subtotal (₹)", "CGST (₹)", "SGST (₹)", "Grand Total (₹)"])
    for col_idx in range(1, 8):
        cell = ws_invoices.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    row_num = 2
    for inv in invoices:
        ws_invoices.append([
            inv.invoice_number,
            inv.created_at.strftime('%Y-%m-%d %H:%M'),
            inv.customer.user.username,
            float(inv.subtotal),
            float(inv.cgst_total),
            float(inv.sgst_total),
            float(inv.grand_total)
        ])
        for col_idx in range(1, 8):
            cell = ws_invoices.cell(row=row_num, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            cell.alignment = align_right if col_idx >= 4 else (align_center if col_idx <= 2 else align_left)
            if row_num % 2 == 1:
                cell.fill = fill_zebra
        row_num += 1
        
    # Auto-adjust column widths
    for ws in [ws_summary, ws_invoices]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Save the output file
    media_dir = os.path.join(settings.BASE_DIR, 'media', 'reports')
    os.makedirs(media_dir, exist_ok=True)
    
    file_path = os.path.join(media_dir, f"gst_summary_{year}_{month:02d}.xlsx")
    wb.save(file_path)
    
    # Return file path
    return file_path


import logging
logger = logging.getLogger(__name__)

@shared_task
def send_whatsapp_notification_task(khata_profile_id, message_type, context_data=None):
    """
    Background Celery task to send formatted WhatsApp notifications to customers
    and log the transaction status in WhatsAppLog.
    """
    from store_app.models import KhataProfile, WhatsAppLog
    from store_app.utils.whatsapp_helpers import send_whatsapp_message

    try:
        profile = KhataProfile.objects.select_related('user').get(id=khata_profile_id)
    except KhataProfile.DoesNotExist:
        logger.error(f"KhataProfile with ID {khata_profile_id} not found for WhatsApp notification.")
        return False

    customer_name = profile.user.first_name or profile.user.username
    phone_number = profile.user.phone_number or ""
    
    if not phone_number:
        WhatsAppLog.objects.create(
            khata_profile=profile,
            message_type=message_type,
            phone_number="N/A",
            message_body="[Failed: No phone number configured on customer profile]",
            status='FAILED',
            error_message="Phone number is empty on User model"
        )
        logger.warning(f"Failed to send WhatsApp message to KhataProfile {khata_profile_id}: No phone number.")
        return False

    balance = float(profile.current_balance)
    body = ""

    if message_type == 'TRANSACTION_ALERT':
        tx_type = context_data.get('transaction_type', 'DEBIT') if context_data else 'DEBIT'
        amount = context_data.get('amount', 0.0) if context_data else 0.0
        desc = context_data.get('description', '') if context_data else ''
        
        # CREDIT means checkout on credit (increases debt), DEBIT means paid balance (decreases debt)
        label = "Checkout on Credit" if tx_type == 'CREDIT' else "Ledger Payment"
        
        body = (
            f"Namaste {customer_name}!\n\n"
            f"A ledger entry has been logged at Shivam Kirana Store:\n"
            f"• Transaction: {label}\n"
            f"• Amount: ₹{amount:.2f}\n"
            f"• Note: {desc}\n\n"
            f"Your running outstanding balance is: ₹{balance:.2f}.\n\n"
            f"Thank you for shopping with us!"
        )
    elif message_type == 'PAYMENT_REMINDER':
        checkout_link = context_data.get('checkout_link') if context_data else None
        if not checkout_link:
            checkout_link = "http://localhost:5174/dashboard/khata"
            
        body = (
            f"Namaste {customer_name}!\n\n"
            f"This is a friendly payment reminder from Shivam Kirana Store.\n"
            f"Your current outstanding balance is: ₹{balance:.2f}.\n\n"
            f"You can settle your balance online using this secure checkout page:\n"
            f"{checkout_link}\n\n"
            f"Alternatively, you may pay cash at the store counter. Thank you!"
        )
    elif message_type == 'STATEMENT':
        total_credit = float(profile.total_credit)
        total_paid = float(profile.total_paid)
        body = (
            f"Namaste {customer_name}!\n\n"
            f"Here is your ledger statement summary from Shivam Kirana Store:\n"
            f"• Outstanding Balance: ₹{balance:.2f}\n"
            f"• Lifetime Purchases: ₹{total_credit:.2f}\n"
            f"• Lifetime Paid: ₹{total_paid:.2f}\n\n"
            f"Please visit [http://localhost:5174/dashboard/khata] to view your full transaction history.\n\n"
            f"Thank you!"
        )
    else:
        body = context_data.get('body', '') if context_data else ''
        if not body:
            body = f"Namaste {customer_name}, your account balance is ₹{balance:.2f}."

    # Create the log in PENDING status
    log_entry = WhatsAppLog.objects.create(
        khata_profile=profile,
        message_type=message_type,
        phone_number=phone_number,
        message_body=body,
        status='PENDING'
    )

    # Dispatch message
    success, err = send_whatsapp_message(phone_number, body)

    if success:
        log_entry.status = 'SENT'
    else:
        log_entry.status = 'FAILED'
        log_entry.error_message = err

    log_entry.save()
    return success


@shared_task
def scan_and_alert_expiring_products_task():
    """
    Periodic Celery task that scans all products and expiry batches for items
    that are expired or expiring within the next 7 days.
    Creates EXPIRY_ALERT notifications for all admin users.
    Can also be triggered manually via the admin API.
    Returns a summary dict with counts.
    """
    from datetime import timedelta
    from django.utils import timezone
    from store_app.models import Product, ExpiryBatch, Notification, CustomUser

    today = timezone.now().date()
    alert_threshold = today + timedelta(days=7)

    # --- Scan product-level expiry dates ---
    expired_products = Product.objects.filter(
        expiry_date__isnull=False,
        expiry_date__lt=today
    ).select_related()

    expiring_soon_products = Product.objects.filter(
        expiry_date__isnull=False,
        expiry_date__gte=today,
        expiry_date__lte=alert_threshold
    ).select_related()

    # --- Scan batch-level expiry dates ---
    expired_batches = ExpiryBatch.objects.filter(
        expiry_date__lt=today
    ).select_related('product')

    expiring_soon_batches = ExpiryBatch.objects.filter(
        expiry_date__gte=today,
        expiry_date__lte=alert_threshold
    ).select_related('product')

    # --- Build admin notification messages ---
    admin_users = list(CustomUser.objects.filter(role='ADMIN'))
    notifications_created = 0

    # Product-level expired items
    for product in expired_products:
        days_ago = (today - product.expiry_date).days
        message = (
            f"[EXPIRED] {product.name} expired {days_ago} day(s) ago "
            f"(Exp: {product.expiry_date.strftime('%d %b %Y')}). "
            f"Stock: {product.stock_quantity} units. Please remove from shelves."
        )
        for admin in admin_users:
            Notification.objects.get_or_create(
                user=admin,
                notification_type='EXPIRY_ALERT',
                message=message,
                defaults={'is_read': False}
            )
        notifications_created += len(admin_users)

    # Product-level expiring soon
    for product in expiring_soon_products:
        days_left = (product.expiry_date - today).days
        message = (
            f"[EXPIRING SOON] {product.name} expires in {days_left} day(s) "
            f"(Exp: {product.expiry_date.strftime('%d %b %Y')}). "
            f"Stock: {product.stock_quantity} units. Consider discounting or returning to supplier."
        )
        for admin in admin_users:
            Notification.objects.get_or_create(
                user=admin,
                notification_type='EXPIRY_ALERT',
                message=message,
                defaults={'is_read': False}
            )
        notifications_created += len(admin_users)

    # Batch-level expired items
    for batch in expired_batches:
        days_ago = (today - batch.expiry_date).days
        batch_label = f"Batch {batch.batch_number}" if batch.batch_number else "Unnamed Batch"
        message = (
            f"[BATCH EXPIRED] {batch.product.name} -- {batch_label} "
            f"expired {days_ago} day(s) ago (Exp: {batch.expiry_date.strftime('%d %b %Y')}). "
            f"Qty: {batch.quantity} units. Immediate action required."
        )
        for admin in admin_users:
            Notification.objects.get_or_create(
                user=admin,
                notification_type='EXPIRY_ALERT',
                message=message,
                defaults={'is_read': False}
            )
        notifications_created += len(admin_users)

    # Batch-level expiring soon
    for batch in expiring_soon_batches:
        days_left = (batch.expiry_date - today).days
        batch_label = f"Batch {batch.batch_number}" if batch.batch_number else "Unnamed Batch"
        message = (
            f"[BATCH EXPIRING SOON] {batch.product.name} -- {batch_label} "
            f"expires in {days_left} day(s) (Exp: {batch.expiry_date.strftime('%d %b %Y')}). "
            f"Qty: {batch.quantity} units."
        )
        for admin in admin_users:
            Notification.objects.get_or_create(
                user=admin,
                notification_type='EXPIRY_ALERT',
                message=message,
                defaults={'is_read': False}
            )
        notifications_created += len(admin_users)

    summary = {
        'expired_products': expired_products.count(),
        'expiring_soon_products': expiring_soon_products.count(),
        'expired_batches': expired_batches.count(),
        'expiring_soon_batches': expiring_soon_batches.count(),
        'notifications_created': notifications_created,
        'scan_date': str(today),
    }
    logger.info(f"[EXPIRY SCAN] Completed: {summary}")
    return summary


@shared_task
def send_otp_email_task(user_id):
    import traceback
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    from django.conf import settings
    from store_app.models import CustomUser

    try:
        user = CustomUser.objects.get(id=user_id)
    except CustomUser.DoesNotExist:
        logger.error(f"[OTP EMAIL] User with ID {user_id} not found for OTP email.")
        return False

    if not user.otp_code:
        logger.error(f"[OTP EMAIL] User {user.username} does not have an active OTP code generated.")
        return False

    # Log the email backend so we can diagnose config issues on Render
    logger.info(
        f"[OTP EMAIL] Preparing to send OTP to {user.email} | "
        f"Backend={settings.EMAIL_BACKEND} | "
        f"Host={settings.EMAIL_HOST}:{settings.EMAIL_PORT} | "
        f"From={settings.DEFAULT_FROM_EMAIL} | "
        f"User={settings.EMAIL_HOST_USER or '(NOT SET)'}"
    )

    subject = f"Verification Code: {user.otp_code} - Shivam Kirana Store"

    from datetime import datetime
    context = {
        'user': user,
        'customer_name': user.username,
        'otp_code': user.otp_code,
        'expiry_minutes': 10,
        'support_email': getattr(settings, 'EMAIL_HOST_USER', 'shivamkiranastoreofficial@gmail.com'),
        'current_year': datetime.now().year
    }

    try:
        html_message = render_to_string('emails/otp_email.html', context)
    except Exception as e:
        logger.warning(f"[OTP EMAIL] Could not render HTML template: {str(e)}. Falling back to plain text.")
        html_message = None

    plain_message = (
        f"Namaste {user.username}!\n\n"
        f"Thank you for registering at Shivam Kirana Store.\n"
        f"Your 6-digit verification code is: {user.otp_code}\n\n"
        f"This code is valid for 10 minutes.\n"
        f"If you did not register this account, please ignore this email.\n"
    )

    try:
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            html_message=html_message,
            fail_silently=False
        )
        logger.info(f"[OTP EMAIL] SUCCESS: OTP email sent successfully to {user.email}")
        return True
    except Exception as e:
        logger.error(
            f"[OTP EMAIL] FAILED: Failed to send OTP email to {user.email}.\n"
            f"Error type: {type(e).__name__}\n"
            f"Error detail: {str(e)}\n"
            f"Traceback:\n{traceback.format_exc()}"
        )
        return False
