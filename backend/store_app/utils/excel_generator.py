import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone

def generate_excel_response(sheet_name, title, table_headers, table_data, summary_data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name='Calibri', size=16, bold=True, color='10B981')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    body_font = Font(name='Calibri', size=11, bold=False)
    bold_body_font = Font(name='Calibri', size=11, bold=True)
    
    header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
    summary_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    
    thin_border_side = Side(border_style='thin', color='CBD5E1')
    border_style = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Title Block
    ws['A1'] = "SHIVAM KIRANA STORE"
    ws['A1'].font = title_font
    
    ws['A2'] = f"{title} (Generated: {timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')})"
    ws['A2'].font = Font(name='Calibri', size=11, italic=True, color='475569')
    
    current_row = 4
    
    # Summary Block
    if summary_data:
        ws.cell(row=current_row, column=1, value="Summary Indicators").font = Font(name='Calibri', size=12, bold=True, color='0F172A')
        current_row += 1
        for key, val in summary_data.items():
            ws.cell(row=current_row, column=1, value=key).font = bold_body_font
            ws.cell(row=current_row, column=1).fill = summary_fill
            ws.cell(row=current_row, column=1).border = border_style
            
            ws.cell(row=current_row, column=2, value=val).font = body_font
            ws.cell(row=current_row, column=2).fill = summary_fill
            ws.cell(row=current_row, column=2).border = border_style
            current_row += 1
        current_row += 2 # Spacer
        
    # Table Headers
    for col_idx, h in enumerate(table_headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center' if 'Date' in h or 'ID' in h else 'left')
        cell.border = border_style
        
    current_row += 1
    
    # Table Rows
    start_row = current_row
    for row_idx, row in enumerate(table_data):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = body_font
            cell.border = border_style
            
            # Formatting and alignments
            if isinstance(val, (int, float, Decimal := type(0.1))):
                cell.alignment = Alignment(horizontal='right')
                if '₹' in table_headers[col_idx-1] or 'Price' in table_headers[col_idx-1] or 'Amount' in table_headers[col_idx-1] or 'Balance' in table_headers[col_idx-1] or 'Due' in table_headers[col_idx-1] or 'Valuation' in table_headers[col_idx-1]:
                    cell.number_format = '₹#,##0.00'
            elif '-' in str(val) and len(str(val)) <= 10:  # Simple date guess
                cell.alignment = Alignment(horizontal='center')
                
            # Zebra striping
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
                
        current_row += 1
        
    # Auto column width adjustment
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Avoid using title lines (row 1 & 2) for length calculation
            if cell.row in [1, 2]:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
